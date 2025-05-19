import os
import openpyxl
import spacy
from openai import AsyncOpenAI
import time
import json
import chardet
import asyncio
from docx import Document
from tqdm import tqdm
import pandas as pd
from tqdm.asyncio import tqdm_asyncio

openai = AsyncOpenAI(
    api_key="sk-db3f839bc51e459dae3aab49d1a779e2",
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
)

# openai = AsyncOpenAI(
#     api_key="sk-cLHG0jRuBeFDE49617b9T3BLBkFJe5b79d2bDefD4Db7b9fa",
#     base_url="https://c-z0-api-01.hash070.com/v1",
# )

nlp = spacy.load("zh_core_web_sm")


def load_config():
    current_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    config_file = os.path.join(current_dir, "config.json")

    with open(config_file, "rb") as f:
        encoding = chardet.detect(f.read())["encoding"]

    with open(config_file, "r", encoding=encoding) as f:
        return json.load(f)

async def replace_character(scenarios, character_dict):
    system_prompt = (
        "你将收到一段文本，以及一个包含角色名称与其特征的映射字典。"
        "请识别文本中所有出现的角色名称，以及指代这些角色的代词、描述性名词、外貌称呼或亲属关系描述。"
        "请统一将这些指代或称呼替换为字典中定义的标准角色名称，保持上下文一致，避免遗漏或误替换，并与原文的格式保持一致，在替代的时候不要出现无关的符号。"
        "如果同一角色在不同阶段（如童年、成年）被赋予不同称呼，请根据上下文推理，并统一替换为同一个角色名。"
        "对于包含亲属关系的表达（如“某某的母亲”、“她的父亲”），请结合上下文，明确识别‘某某’指的是谁，"
        "再在字典中查找其母亲（或父亲）是谁，并替换为该角色的标准名称。"
        "例如，如果“白雪公主（儿童）”和“王后（白雪公主生母）”都出现在字典中，且提到“白雪公主（美少女）的母亲”，"
        "应识别出“白雪公主（美少女）”与“白雪公主（儿童）”为同一人，其母亲即“王后（白雪公主生母）”，应替换为后者。"
        "保持文本原有格式和结构，不要添加、修改或省略任何内容，也不要输出 prompt 以外的任何信息。"
    )

    scenario_text = "\n".join(scenarios["Chinese Content"].tolist())

    messages = [
        {"role": "system", "content": system_prompt},
        {
            "role": "user",
            "content": f'角色名称与其特征的映射字典如下：{character_dict}\n\n文本如下：{scenario_text}'
        },
    ]

    resp = await request_with_retry_async(messages)
    return [row.strip() for row in resp.split("\n") if row.strip()]


def replace_keywords(sentence, keyword_dict):
    for key, value in keyword_dict.items():
        if key and value:
            sentence = sentence.replace(key, f'{key}(character features: {value})')
    return sentence


async def request_with_retry_async(
    messages, max_requests=90, cooldown_seconds=60
):
    """异步版本的API请求函数"""
    attempts = 0
    while attempts < max_requests:
        try:
            response = await openai.chat.completions.create(
                model="qwen-plus-latest",
                # model="gpt-4o-mini",
                messages=messages,
            )
            return response.choices[0].message.content.strip()
        except Exception as e:
            print(f"发生错误：{str(e)}")
            await asyncio.sleep(10)
        attempts += 1

    return "请求失败，已达到最大尝试次数"


async def translate_to_english_async(text):
    """异步版本的英文翻译函数"""
    messages = [
        {"role": "system", "content": "You are a helpful assistant."},
        {
            "role": "user",
            "content": f'Translate the following text into English: "{text}". Do not directly translate, but instead translate from a third-person descriptive perspective, and complete the missing subject, predicate, object, attributive, adverbial, and complement in the text. Besides the translated result, do not include any irrelevant content or explanations in your response.',
        },
    ]
    return await request_with_retry_async(messages)


async def translate_to_storyboard_async(text, trigger):
    """异步版本的分镜生成函数"""
    messages = [
        {
            "role": "system",
            "content": "StableDiffusion is a deep learning text-to-image model that supports the generation of new images using keywords to describe the elements to be included or omitted. Now, as a professional StableDiffusion AI drawing keyword generator. You can assist me in generating keywords for my desired image.",
        },
        {"role": "user", "content": f"{trigger}'{text}'"},
    ]
    return await request_with_retry_async(messages)


def read_docx(file_path):
    return [
        paragraph.text
        for paragraph in Document(file_path).paragraphs
        if paragraph.text.strip()
    ]




# 假设这些函数已经定义好
# from your_module import translate_to_english_async, translate_to_storyboard_async, replace_keywords

async def process_text_sentences_async(
    input_file_path,
    output_file_path,
    trigger,
    keyword_dict,
):
    """
    异步处理 CSV 中的文本数据，包含关键词替换、翻译、StableDiffusion关键词生成。
    """

    # 读取 CSV 文件
    dataframe = pd.read_csv(input_file_path)

    # 检查必须列
    if 'Replaced Content' not in dataframe.columns:
        raise ValueError("CSV中缺少'Replaced Content'这一列。")

    # 替换关键词（支持传入自定义的 replace_keywords 函数）
    dataframe['Replaced Content'] = dataframe['Replaced Content'].apply(replace_keywords, args=(keyword_dict,))

    # 异步翻译
    print("🔤 开始翻译内容...")
    translation_tasks = [
        translate_to_english_async(text) for text in dataframe['Replaced Content']
    ]
    dataframe['Translated Content'] = await tqdm_asyncio.gather(*translation_tasks, desc="翻译中", ncols=80)

    # 异步生成分镜脚本
    print("🎨 开始生成分镜脚本...")
    storyboard_tasks = [
        translate_to_storyboard_async(text, trigger) for text in dataframe['Translated Content']
    ]
    dataframe['SD Content'] = await tqdm_asyncio.gather(*storyboard_tasks, desc="生成分镜", ncols=80)

    # 保存结果
    dataframe.to_csv(output_file_path, index=False)
    dataframe.to_excel(output_file_path.replace(".csv", ".xlsx"), index=False)
    print(f"✅ 已保存到 {output_file_path}")


    #     paragraphs = read_docx(input_file_path)
    # except ValueError as e:
    #     print(f"发生错误：{str(e)}")
    #     return
    # print(paragraphs)
    #
    # sentences = []
    # for paragraph in paragraphs:
    #     sentences.extend([sent.text for sent in nlp(paragraph).sents])
    #
    # sentences = merge_short_sentences(sentences, min_sentence_length)

    # original_sentences_dict = {}
    # sheet = workbook.active
    # for idx, sentence in enumerate(sentences, 1):
    #     replaced_sentence, original_sentence = replace_keywords(sentence, keyword_dict)
    #     original_sentences_dict[replaced_sentence] = original_sentence
    #     sheet.cell(row=idx, column=1, value=replaced_sentence)
    #     sheet.cell(row=idx, column=4, value=original_sentence)
    #
    # replaced_sentences = list(original_sentences_dict.keys())


    # # 创建一个进度条计数器
    # translation_progress = tqdm(total=len(replaced_sentences), desc="正在翻译文本")
    # storyboard_progress = tqdm(total=len(replaced_sentences), desc="正在生成分镜脚本")
    #
    # # 使用信号量限制并发请求数
    # sem = asyncio.Semaphore(5)  # 最多5个并发请求
    #
    # async def process_sentence(idx, sentence):
    #     """处理单个句子的翻译和分镜生成"""
    #     async with sem:
    #         # 翻译步骤
    #         translated_text = await translate_to_english_async(sentence.strip())
    #         sheet.cell(row=idx, column=2, value=translated_text)
    #         translation_progress.update(1)
    #
    #         # 分镜生成步骤
    #         storyboard_text = await translate_to_storyboard_async(
    #             translated_text, trigger
    #         )
    #         sheet.cell(row=idx, column=3, value=storyboard_text)
    #         storyboard_progress.update(1)
    #
    # # 创建所有句子的处理任务
    # tasks = [
    #     process_sentence(idx, sentence)
    #     for idx, sentence in enumerate(replaced_sentences, 1)
    # ]
    #
    # # 执行所有任务
    # await asyncio.gather(*tasks)
    #
    # # 关闭进度条
    # translation_progress.close()
    # storyboard_progress.close()
    #
    # # 保存结果
    # workbook.save(output_file_path)


async def main_async():
    """异步版本的主函数"""
    config = load_config()
    print("BADAPPLE")

    role_name = config.get("角色名1", "未指定角色名")
    feature = config.get("特征1", "未指定特征")
    role2_name = config.get("角色名2", "未指定角色名2")
    feature2 = config.get("特征2", "未指定特征2")
    role3_name = config.get("角色名3", "未指定角色名3")
    feature3 = config.get("特征3", "未指定特征3")
    role4_name = config.get("角色名4", "未指定角色名4")
    feature4 = config.get("特征4", "未指定特征4")
    role5_name = config.get("角色名5", "未指定角色名5")
    feature5 = config.get("特征5", "未指定特征5")
    role6_name = config.get("角色名6", "未指定角色名6")
    feature6 = config.get("特征6", "未指定特征6")
    role7_name = config.get("角色名7", "未指定角色名7")
    feature7 = config.get("特征7", "未指定特征7")
    role8_name = config.get("角色名8", "未指定角色名8")
    feature8 = config.get("特征8", "未指定特征8")
    role9_name = config.get("角色名9", "未指定角色名9")
    feature9 = config.get("特征9", "未指定特征9")
    role10_name = config.get("角色名10", "未指定角色名10")
    feature10 = config.get("特征10", "未指定特征10")
    keyword_dict = {
        role_name: feature,
        role2_name: feature2,
        role3_name: feature3,
        role4_name: feature4,
        role5_name: feature5,
        role6_name: feature6,
        role7_name: feature7,
        role8_name: feature8,
        role9_name: feature9,
        role10_name: feature10,
    }

    default_trigger = """""Task: I will give you the theme in natural language. Your task is to imagine a full picture based on that theme and convert it into a high-quality prompt for Stable Diffusion.  

Prompt concept: A prompt describes the content of an image using simple, commonly used English tags separated by English half-width commas (','). Each word or phrase is a tag. The prompt must form a single continuous sentence — do not break it into parts or include labels like 'Prompt:', 'Style:', 'Main subject:' etc., and never use ':' or '.' in the final prompt.  

Prompt requirements: The prompt should include the following elements:
- Main subject (e.g. a girl in a garden), enriched with relevant details depending on the theme.
- For characters, describe facial features like 'beautiful detailed eyes, beautiful detailed lips, extremely detailed eyes and face, long eyelashes' to prevent facial deformities.
- Additional scene or subject-related details.
- Image quality tags such as '(best quality,4k,8k,highres,masterpiece:1.2), ultra-detailed, (realistic,photorealistic,photo-realistic:1.37)' and optionally: HDR, UHD, studio lighting, ultra-fine painting, sharp focus, extreme detail description, professional, vivid colors, bokeh, physically-based rendering.
- Artistic style, color tone, and lighting should also be included in tags.

To prevent trait blending between multiple characters:
- Use 'BREAK' in uppercase between character descriptions.
- Enclose each character in parentheses ().
- Add spatial or layout hints like 'left side', 'in background', 'group of three', etc.

Example for 2 characters:
2people(1boy and 1girl),walking,(street_background:1.3),(looking at viewer), dynamic pose, (masterpiece:1.4, best quality), unity 8k wallpaper, ultra detailed, beautiful and aesthetic, perfect lighting,detailed background, realistic,
==BREAK==
1girl, red long hair and red eyes and (red shirt:1.3),
==BREAK==
1boy, yellow short hair and yellow eyes and (yellow suit:1.3) and hands in pocket,

Output only the final prompt in English, no explanations or additional formatting."

    """

    # default_trigger = """Here, I introduce the concept of Prompts from the StableDiffusion algorithm, also known as hints.
    # The following prompts are used to guide the AI painting model to create images.
    # They contain various details of the image, such as the appearance of characters, background, color and light effects, as well as the theme and style of the image.
    # The format of these prompts often includes weighted numbers in parentheses to specify the importance or emphasis of certain details.
    # For example, "(masterpiece:1.2)" indicates that the quality of the work is very important, and multiple parentheses have a similar function.
    # Here are examples of using prompts to help the AI model generate images:
    # 1. (masterpiece:1.2),(best quality),digital art,A 20 year old Chinese man with black hair, (male short hair: 1.2), green shirt, walking on the road to rural China, ultra wide angle
    # Please use English commas as separators. Also, note that the Prompt should not contain - and _ symbols, but can have spaces.
    #
    # In character attributes, 1girl means you generated a girl, 2girls means you generated two girls.
    # In the generation of Prompts, you need to describe character attributes, theme, appearance, emotion, clothing, posture, viewpoint, action, background using keywords.
    # - You may include `LoRA` tokens such as `<lora:charA:0.8>` for character-specific identity.
    # Please provide a set of prompts that highlight the theme.
    # Note: The prompt cannot exceed 100 words, no need to use natural language description, character attributes need to be highlighted a little bit, for example: {role_name}\({feature}\).
    # If the content contains a character name, add the specified feature as required, if the content does not contain the corresponding character name, then improvise.
    # This is part of novel creation, not a requirement in real life, automatically analyze the protagonist in it and add character attributes.
    #
    # To prevent trait blending between multiple characters:
    # - Use `BREAK` in uppercase between characters to separate attention.
    # - Enclose each character description with `()`.
    # - Add spatial or compositional cues like `left side`, `in background`, `group of three`.
    #
    # Example: (best quality,4k,8k,highres,masterpiece:1.2), ultra-detailed, cinematic lighting, (1girl: long silver hair, red kimono, holding katana, beautiful detailed eyes and lips, sharp gaze, full body), BREAK (1boy: short black hair, armor, green cape, holding shield, looking left), fantasy battlefield background, glowing effects, HDR, epic style, soft sunlight
    # The prompt must be in English, only provide the prompt, no extra information is needed.
    # Here is the content:"""

    # default_trigger = """You are a Stable Diffusion prompt assistant with a strong sense of visual aesthetics.
    #
    # Your job is to transform a concept into a high-quality image prompt compatible with Stable Diffusion 3.5 large. You will only output the positive prompts.
    #
    # ---
    #
    # Prompt Format Rules:
    #
    # 1. Tag structure (ordered by importance):
    #    - (best quality,4k,8k,highres,masterpiece:1.2), ultra-detailed, (realistic,photorealistic,photo-realistic:1.37)
    #    - Subject (e.g., 1girl, knight, dragon), with details like facial features, pose, clothes, expression
    #    - Additional details (background, weather, action, props)
    #    - Colors and lighting (e.g., warm light, blue tones, soft shadows)
    #
    # 2. Use () or (keyword:1.1~1.5) to increase tag weight. Use [] or (keyword:0.9) to reduce importance.
    # 3. Always use English commas `,` as separators.
    # 4. Avoid natural language, colons `:`, or full sentences in the prompt.
    #
    # ---
    #
    # Multi-character Guidelines:
    #
    # To prevent trait blending between multiple characters:
    # - Use `BREAK` in uppercase between characters to separate attention.
    # - Enclose each character description with `()`.
    # - Add spatial or compositional cues like `left side`, `in background`, `group of three`.
    # - You may include `LoRA` tokens such as `<lora:charA:0.8>` for character-specific identity.
    #
    # Please follow the example, and do not limit to the words I give you.
    #
    # Example for 2 characters:
    # (best quality,4k,8k,highres,masterpiece:1.2), ultra-detailed, cinematic lighting, (1girl: long silver hair, red kimono, holding katana, beautiful detailed eyes and lips, sharp gaze, full body), BREAK (1boy: short black hair, armor, green cape, holding shield, looking left), fantasy battlefield background, glowing effects, HDR, epic style, soft sunlight
    #
    # The prompt must be in English, only provide the prompt, no extra information is needed.
    # Here is the content:"""


    # trigger = config.get("引导词", default_trigger)
    trigger = default_trigger
    current_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    input_file_path = os.path.join(current_dir, "txt", "txt.csv")
    output_dir = os.path.join(current_dir, "txt")

    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)

    output_file_path = os.path.join(output_dir, "output2.csv")

    await process_text_sentences_async(
        input_file_path,
        output_file_path,
        trigger,
        keyword_dict,
    )


def main():
    """入口函数，运行异步主函数"""
    asyncio.run(main_async())


if __name__ == "__main__":
    main()
